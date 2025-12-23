from openpyxl import Workbook

from semantic_tester.excel.utils import (
    write_cell_safely,
    get_column_index,
    get_or_add_column,
)
import pandas as pd


def test_write_cell_safely_handles_merged_cells():
    wb = Workbook()
    ws = wb.active

    # 合并 A1:C1，并向中间单元格写值，应写到左上角 A1
    ws.merge_cells("A1:C1")

    write_cell_safely(ws, 1, 2, "merged")  # B1

    assert ws["A1"].value == "merged"
    # 未合并单元格直接写入
    write_cell_safely(ws, 2, 1, "normal")
    assert ws["A2"].value == "normal"


def test_get_column_index_and_get_or_add_column():
    cols = ["A", "B", "C"]

    # 按序号
    assert get_column_index(cols, "1") == 0
    assert get_column_index(cols, "3") == 2
    assert get_column_index(cols, "4") == -1

    # 按列名
    assert get_column_index(cols, "B") == 1
    assert get_column_index(cols, "X") == -1

    df = pd.DataFrame({"A": [1], "B": [2]})
    col_names = list(df.columns)

    # 选择已有列
    idx = get_or_add_column(df, col_names, "1")  # -> A
    assert idx == 0

    # 无效序号时新增列，列名就是原始输入
    idx_new = get_or_add_column(df, col_names, "5")
    assert col_names[idx_new] == "5"
    assert "5" in df.columns

    # 通过列名新增
    idx_name = get_or_add_column(df, col_names, "result")
    assert col_names[idx_name] == "result"
    assert "result" in df.columns
